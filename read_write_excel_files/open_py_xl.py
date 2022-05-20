""" 
    using openPyXl
"""
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, colors
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
import datetime as dt
import openpyxl
import pandas as pd
import excel

# -----------------------------------------------------------------------------------------------------------------
#       reading an excel file wiht openpyxl
# Open the workbook to read cell values.
# The file is automatically closed again after loading the data.
book = openpyxl.load_workbook("xl/stores.xlsx", data_only=True)
# Get a worksheet object by name or index (0-based)
# sheet = book["2019"]
sheet = book.worksheets[0]
# Get a list with all sheet names
# print(book.sheetnames)
# Out[4]: ['2019', '2020', '2019-2020']
# Loop through the sheet objects.
# Instead of "name", openpyxl uses "title".
for i in book.worksheets:
    print(i.title)
# 2019
# 2020
# 2019-2020
# Getting the dimensions, the used range of the sheet
print(sheet.max_row, sheet.max_column)
# Out[6]: (8, 6)

# Read the value of a single cell
# using "A1" notation and using cell indices (1-based)
# print(sheet["B6"].value)
a = sheet.cell(row=6, column=2).value
print(a)
# Out[7]: 'Boston'

# Read in a range of cell values by using our excel module
data = excel.read(book["2019"], (2, 2), (8, 6))
print(data[:2])  # Print the first two rows
# Out[8]: [['Store', 'Employees', 'Manager', 'Since', 'Flagship'],
#          ['New York', 10, 'Sarah', datetime.datetime(2018, 7, 20, 0, 0), False]]

# ---------------------------------------------------------------------------------------------------------
#       Writing with OpenPyXL
# OpenPyXL builds the Excel file in memory and writes out the file once you call the
# save method. The following code produces the file as shown in Figure 8-1:
# import openpyxl
# Instantiate a workbook
book = openpyxl.Workbook()
# Get the first sheet and give it a name
sheet = book.active
sheet.title = "Sheet1"
# Writing individual cells using A1 notation
# and cell indices (1-based)
sheet["A1"].value = "Hello 1"
sheet.cell(row=2, column=1, value="Hello 2")


# Formatting: fill color, alignment, border and font
font_format = Font(color="FF0000", bold=True)
thin = Side(border_style="thin", color="FF0000")
sheet["A3"].value = "Hello 3"
sheet["A3"].font = font_format
sheet["A3"].border = Border(top=thin, left=thin,
                            right=thin, bottom=thin)
sheet["A3"].alignment = Alignment(horizontal="center")
sheet["A3"].fill = PatternFill(fgColor="FFFF00", fill_type="solid")
# Number formatting (using Excel's formatting strings)
sheet["A4"].value = 3.3333
sheet["A4"].number_format = "0.00"
# Date formatting (using Excel's formatting strings)
sheet["A5"].value = dt.date(2016, 10, 13)
sheet["A5"].number_format = "mm/dd/yy"
# Formula: you must use the English name of the formula
# with commas as delimiters
sheet["A6"].value = "=SUM(A4, 2)"
# Image
sheet.add_image(Image("images/python.png"), "C1")
# Two-dimensional list (we're using our excel module)
data = [[None, "North", "South"],
        ["Last Year", 2, 5],
        ["This Year", 3, 6]]
excel.write(sheet, data, "A10")
# Chart
chart = BarChart()
chart.type = "col"
chart.title = "Sales Per Region"
chart.x_axis.title = "Regions"
chart.y_axis.title = "Sales"
chart_data = Reference(sheet, min_row=11, min_col=1,
                       max_row=12, max_col=3)
chart_categories = Reference(sheet, min_row=10, min_col=2,
                             max_row=10, max_col=3)
# from_rows interprets the data in the same way
# as if you would add a chart manually in Excel
chart.add_data(chart_data, titles_from_data=True, from_rows=True)
chart.set_categories(chart_categories)
sheet.add_chart(chart, "A15")
# Saving the workbook creates the file on disk
book.save("openpyxl.xlsx")
# --------------------------------------------------------------------------------------------------------
# If you want to write an Excel template file, 
# you’ll need to set the template attribute to
# True before saving it:
book = openpyxl.Workbook()
sheet = book.active
sheet["A1"].value = "This is a template"
book.template = True
book.save("template.xltx")
# ---------------------------------------------------------------------------------------------------------
# OpenPyXL will either change them or drop them altogether. For example, as of
# v3.0.5, OpenPyXL will rename charts and drop their title. Here is a simple editing
# example:
# Read the stores.xlsx file, change a cell
    # and store it under a new location/name.
book = openpyxl.load_workbook("xl/stores.xlsx")
book["2019"]["A1"].value = "modified"
book.save("stores_edited.xlsx")
# ---------------------------------------------------------------------------------------------------------

# If you want to write an xlsm file, OpenPyXL has to work off an existing file that you
# need to load with the keep_vba parameter set to True:
book = openpyxl.load_workbook("xl/macro.xlsm", keep_vba=True)
book["Sheet1"]["A1"].value = "Click the button!"
book.save("macro_openpyxl.xlsm")
# The button in the example file is calling a macro that shows a message box.
